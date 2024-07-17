#Programa PYTHON
#Sistema: AGS CIS/AUDIT
#Definicion de Vistas para el Sistema BCP del Proyecto DEFCON5
#Programado por Marco A. Villalobos Michelson
#==============================================================

#========
# Indice
#========

#  Framework

#  Lista FrameWork *
#  Crea FrameWork *
#  Borra  el Framework *
#  Lista Modulo Modelo y Objetivos de Control asociados *
#  Crea Modulo Modelo en el Framework *
#  Borra Modulo Modelo en el Framework *
#  Crea OC Modelo en el Framework *
#  Borra OC Modelo en el Framework *
#  Modifica el OC Modelo en el Framework *

# Auditorias

# Lista Modulo Modelo y Objetivos de Control asociados







from django.shortcuts import render
from django.contrib import messages

# Modelos/Entidades

#from .models import Proceso, SubProceso, LogAut, Recursos, Tipo_RR, Gestor, Escenarios, Amenazas, Estrategias, Tipo_Impacto, Nivel_Impacto

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User, PermissionsMixin
from django.contrib.auth.hashers import make_password

from django.views import generic
from django.shortcuts import get_object_or_404
from django.views.generic.base import TemplateView

from django.http import HttpResponseRedirect
from django.urls import reverse 
import datetime

from django.contrib.auth.decorators import permission_required

#from django.contrib.sites.models import Site
from django.core.mail import send_mail


#Variables Globales

# Fin variables Globales


def index(request):
    """
    Función de vista para la página inicio del sitio.
    """

    #usr=request.user
    #if not usr.is_authenticated:
        #return HttpResponseRedirect(reverse('error-sesion-mgm', args=[500] ))

    # Genera contadores de algunos de los objetos principales
    #num_proc=Proceso.objects.all().count()
    #num_sproc=SubProceso.objects.all().count()
    #num_proced=Procedimientos.objects.all().count()
   
    # Numero de visitas a esta view, como está contado en la variable de sesión.
    num_visits = request.session.get('num_visits', 0)
    request.session['num_visits'] = num_visits + 1

    #defcon = get_object_or_404(Parametros_G , pk=1)

    # Renderiza la plantilla HTML index.html con los datos en la variable contexto
    return render(request, 'index.html')
        
#********************************************************************************
#******************************** FrameWork *************************************
#********************************************************************************
from .models import Framework, Parametros_G, Modulo_Mod, Obj_Ries_Mod


#*******************
#  Lista FrameWork *
#*******************

def Lista_Framew(request):

    print('Entra a Lista Frame')
    lista_framework = Framework.objects.all()
    
    #return(request, 'CIS/Frameworks/framework_list', context={'lista_procesos':lista_procesos})
    return render(request, 'CIS/Frameworks/framework_list.html',
                  context={'lista_framework':lista_framework})
    

#******************
#  Crea FrameWork *
#******************
from .forms import Crea_FrameForm

def Crea_Frame(request):
    """
    Crea un FrameWork
    """
    print('*** Entra a Crea Frame ***')

    # Determinacion de Codigo a asignar.
    cod = get_object_or_404(Parametros_G, pk = 1)
    if cod.valor_2 < 10:
        codigo='0'+str(cod.valor_2)
    else:
        codigo=str(cod.valor_2)
    cod.valor_2=cod.valor_2+1
    cod.save()

    # Ingreso y Registro de Framework
    if request.method == 'POST':

        form = Crea_FrameForm(request.POST)

        if form.is_valid():

            # Crea en registro
            frame = Framework()

            frame.numero = codigo
            #frame.codigo=form.cleaned_data['codigo_frame']
            nombre = form.cleaned_data['nombre']
            frame.fw_nombre = nombre
            version = form.cleaned_data['version']
            frame.fw_version = version

            frame.fw_descrip = form.cleaned_data['descripcion']
            frame.fw_inst = form.cleaned_data['institucion']
            
            frame.save()

            # Crea Modulo Raiz
            mod = frame.fw_modulos = Modulo_Mod()

            mod.pk_frw = frame.pk
            mod.pk_padre=frame.pk
            mod.numero = codigo
            mod.item = '0.- Root'
            mod.nombre =nombre+'/'+version
            mod.framework = frame 
            mod.save()

            #frame.save()


            return HttpResponseRedirect(reverse('Lista-Framework'))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_FrameForm()
        return render(request, 'CIS/Frameworks/crea_framework.html', {'form':form})


#***********************
#  Borra  el Framework *
#***********************
from .forms import Crea_ModuloForm

def Borra_Framework(request, pk):

    print('Entra a Borra Framework pk=', pk)

    frame = get_object_or_404(Framework, pk=pk)
    
    if frame.nro_modulos == 0 :
        frame.delete()
        
    return HttpResponseRedirect(reverse('Lista-Framework'))



#***********************************************************************
#  Lista Modulo Modelo y Objetivos de Control del Framework  asociados *
#***********************************************************************

def Lista_Mod_M(request, pk):

    print('Entra a Lista Modulo Modelo')

    #selecciona todos los Modulos de ese Framework
    framework = get_object_or_404(Framework, pk=pk)
    lista_modulos = Modulo_Mod.objects.filter(pk_frw=pk)
    
    #Crea una lista de OC por cada Modulo del Framework
    matriz_oc=[]
    for mod in lista_modulos:
        
        lista_oc= Obj_Ries_Mod.objects.filter(pk_padre=mod.pk)
        matriz_oc.append({'modulo':mod.pk, 'lista_oc':lista_oc})
        #mod.abierto=False   #Cierra las listas de OC de cada modulo en pagina 
        mod.save() 

    #print('matriz oc =', matriz_oc)

    return render(request, 'CIS/Frameworks/modulos_list.html',
                  context={'lista_modulos':lista_modulos,'matriz_oc':matriz_oc, 'framework':framework})



#**************************************
#  Crea Modulo Modelo en el Framework *
#**************************************
from .forms import Crea_ModuloForm

def Crea_Mod_Mod(request, pk):

    print('Entra a Crea Modulo Modelo pk=', pk)

    mod = get_object_or_404(Modulo_Mod, pk=pk)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    frame=mod.framework
    
    # Determinacion de Codigo a asignar.
    cod = mod.nro_submodulos
    mod.nro_submodulos = cod+1
    mod.save()
    
    if cod < 10:
        codigo='0'+str(cod)
    else:
        codigo=str(cod)

    codigo=mod.numero+codigo

    #Crea Modulo Modelo
    if request.method == 'POST':

        form = Crea_ModuloForm(request.POST)

        if form.is_valid():

            # Crea en registro
            mod_hijo=Modulo_Mod()

            mod_hijo.pk_padre=pk
            mod_hijo.pk_frw=mod.pk_frw
            mod_hijo.numero=codigo
            mod_hijo.item=form.cleaned_data['item']
            mod_hijo.nombre=form.cleaned_data['nombre']
            mod_hijo.framework=frame

            #Incrementa el nro. de Modulos del FrameWork
            frame.nro_modulos=frame.nro_modulos+1
            mod_hijo.save()
            frame.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_ModuloForm()
        return render(request, 'CIS/Frameworks/crea_modulo_m.html', {'form':form,'frame':frame})


#***************************************
#  Borra Modulo Modelo en el Framework *
#***************************************
from .forms import Crea_ModuloForm

def Borra_Mod_Mod(request, pk):

    print('Entra a Borra Modulo Modelo pk=', pk)

    mod = get_object_or_404(Modulo_Mod, pk=pk)
    mod_p = get_object_or_404(Modulo_Mod, pk=mod.pk_padre)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    frame=mod.framework

    if mod.nro_objetivos == 0 and mod.nro_submodulos == 0 :
        frame.nro_modulos=frame.nro_modulos-1

        mod_p.nro_submodulos=mod_p.nro_submodulos-1

        frame.save()
        mod_p.save()
        mod.delete()

    return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))



#**********************************
#  Crea OC Modelo en el Framework *
#**********************************
from .forms import Crea_OCForm

def Crea_OC_Mod(request, pk):

    print('Entra a Crea OC Modelo pk=', pk)

    mod = get_object_or_404(Modulo_Mod, pk=pk)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    frame=mod.framework
    
    # Determinacion de Codigo a asignar.
    cod = mod.nro_objetivos
    mod.nro_objetivos = cod+1
    mod.save()
    
    if cod < 10:
        codigo='0'+str(cod)
    else:
        codigo=str(cod)

    codigo=mod.numero+codigo

    #Crea Modulo Modelo
    if request.method == 'POST':

        form = Crea_OCForm(request.POST)

        if form.is_valid():

            # Crea en registro
            oc=Obj_Ries_Mod()

            oc.pk_padre=pk
            oc.numero=codigo
            oc.modulo=mod
            oc.obj_item=form.cleaned_data['item']
            oc.obj_nombre=form.cleaned_data['nombre']
            oc.obj_descripcion=form.cleaned_data['descripcion']
            oc.save()
            #frame.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_OCForm()
        return render(request, 'CIS/Frameworks/crea_oc_m.html', {'frame':frame, 'mod':mod, 'form':form})


#***********************************
#  Borra OC Modelo en el Framework *
#***********************************
from .forms import Crea_ModuloForm

def Borra_OC_Mod(request, pk):

    print('Entra a Borra Modulo Modelo pk=', pk)

    oc = get_object_or_404(Obj_Ries_Mod, pk=pk)
    mod_p = get_object_or_404(Modulo_Mod, pk=oc.pk_padre)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    #frame=mod.framework

    if mod_p.nro_objetivos == 0 :
        mod_p.nro_objetivos=mod_p.nro_objetivos-1
        oc.delete()
        mod_p.save()

    return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(mod_p.pk_frw)]))


#*****************************************
#  Modifica el OC Modelo en el Framework *
#*****************************************
from .forms import Crea_OCForm

def Mod_OC_Mod(request, pk):

    print('Entra a Modfica el  OC Modelo pk=', pk)

    oc = get_object_or_404(Obj_Ries_Mod, pk=pk)
    mod= get_object_or_404(Modulo_Mod, pk=oc.pk_padre)
    frame=get_object_or_404(Framework, pk=mod.pk_frw)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
        
    #Modifica el  Modulo Modelo
    if request.method == 'POST':

        form = Crea_OCForm(request.POST)

        if form.is_valid():

            # Crea en registro
            
            oc.obj_item=form.cleaned_data['item']
            oc.obj_nombre=form.cleaned_data['nombre']
            oc.obj_descripcion=form.cleaned_data['descripcion']
            oc.save()
            #frame.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_OCForm(initial={'item':oc.obj_item,
                                 'nombre':oc.obj_nombre,
                                 'descripcion':oc.obj_descripcion})
        
        return render(request, 'CIS/Frameworks/modif_oc_m.html', {'frame':frame, 'mod':mod, 'form':form})


#********************************************************************************
#******************************** Auditorias ************************************
#********************************************************************************
from .models import Auditorias, Modulo, Obj_Ries


#********************
#  Lista Auditorias *
#********************

def Lista_Auditorias(request):

    print('Entra a Lista Frame')
    lista_auditorias = Auditorias.objects.all()
    
    return render(request, 'CIS/Auditorias/auditorias_list.html',
                  context={'lista_auditorias':lista_auditorias})



#******************
#  Crea Auditoria *
#******************
from .forms import Crea_AuditoriaForm

def Crea_Auditoria(request):
    """
    Crea una Auditoria
    """
    print('*** Entra a Crea Auditoria ***')

    # Determinacion de Codigo a asignar.
    cod = get_object_or_404(Parametros_G, pk = 2)
    if cod.valor_2 < 10:
        codigo='0'+str(cod.valor_2)
    else:
        codigo=str(cod.valor_2)
    cod.valor_2=cod.valor_2+1
    cod.save()

    # Ingreso y Registro de Framework
    if request.method == 'POST':

        form = Crea_AuditoriaForm(request.POST)

        if form.is_valid():

            # Crea en registro
            aud = Auditorias()

            aud.numero = codigo
            aud.cod_aud=form.cleaned_data['codigo']

            nombre = form.cleaned_data['nombre']
            aud.nombre = nombre

            objetivo = form.cleaned_data['objetivo']
            aud.objetivo = objetivo

            
            auditor_ejec= form.cleaned_data['auditor_ejec']
            aud.auditor_ejec=auditor_ejec
            auditor_sup = form.cleaned_data['auditor_sup']
            aud.auditor_sup=auditor_sup

            aud.fec_ini=form.cleaned_data['fecha_ini']
            aud.fec_ter=form.cleaned_data['fecha_ter']
            
            framework=form.cleaned_data['framework']
            aud.framework=framework
            
            aud.save()

            # Crea Modulo Raiz
            mod_raiz = aud.fw_modulos = Modulo()

            mod_raiz.pk_aud = aud.pk
            mod_raiz.pk_padre=aud.pk
            mod_raiz.numero = codigo
            mod_raiz.item = '0.- Root'
            mod_raiz.nombre =nombre
            mod_raiz.auditoria = aud
            mod_raiz.save()

            #Define Modulos y OC desde el Framework

            if framework:

                aud.nro_modulos=framework.nro_modulos
                pk_frame=framework.pk
                print('*** Identifica  Framework a Copiar***', pk_frame)

                modulos_del_framework=Modulo_Mod.objects.filter(pk_frw=pk_frame)
                
                pk_padre=mod_raiz.pk_aud  #asigna el pk del modulo Raiz

                for mod_modelo in modulos_del_framework:

                    if  mod_modelo.item != '0.- Root':

                        mod_nuevo =Modulo()

                        mod_nuevo.pk_aud=aud.pk #Asigna el Pk de la Auditoria
                        mod_nuevo.pk_padre=pk_padre #Asigna el pk del padre
                        

                        mod_nuevo.numero=mod_modelo.numero
                        mod_nuevo.item=mod_modelo.item
                        mod_nuevo.nombre=mod_modelo.nombre
                        mod_nuevo.auditoria=aud
                        mod_nuevo.nro_submodulos=mod_modelo.nro_submodulos
                        mod_nuevo.nro_objetivos=mod_modelo.nro_objetivos
                        print('Crea Modulo Nuevo:', mod_nuevo.numero,'-', mod_nuevo.item, '-',mod_nuevo.nombre)
                        mod_nuevo.save()

                        pk_padre=mod_nuevo.pk #Asigna el pk del nuevo registra al padre para ser asignado al siguiente

                        #Identifica OC del Framework a Copiar
                        pk_padre_modelo= mod_modelo.pk
                        print('== Pk padre Modelo=', pk_padre_modelo)
                        oc_del_framework =Obj_Ries_Mod.objects.filter(pk_padre=pk_padre_modelo)
                        print(oc_del_framework)

                        if oc_del_framework:

                            for oc_modelo in oc_del_framework:
                            
                                oc_nuevo=Obj_Ries()

                                oc_nuevo.pk_padre=mod_nuevo.pk
                                oc_nuevo.numero  = oc_modelo.numero
                                oc_nuevo.obj_item= oc_modelo.obj_item
                                oc_nuevo.obj_nombre= oc_modelo.obj_nombre
                                oc_nuevo.obj_descripcion= oc_modelo.obj_descripcion
                                #print('mod_nuevo = ', mod_nuevo)
                                oc_nuevo.modulo = mod_nuevo
                                oc_nuevo.save()
                                print('Copia OC del FRamework :', oc_nuevo.obj_item, '-', oc_nuevo.obj_nombre )

                        else:
                            print('No se detectan OC')

                    
            
            return HttpResponseRedirect(reverse('Lista-Auditorias'))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)
            return HttpResponseRedirect(reverse('Lista-Auditorias'))
            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_AuditoriaForm()
        return render(request, 'CIS/Auditorias/crea_auditoria.html', {'form':form})


#********************************************************
#  Lista Modulo Modelo y Objetivos de Control asociados *
#********************************************************

def Lista_Mod(request, pk):

    print('Entra a Lista Modulo ')

    #selecciona todos los Modulos de esa Auditoria
    auditorias = get_object_or_404(Auditorias, pk=pk)
    lista_modulos = Modulo.objects.filter(pk_aud=pk)
    print('Lista de Modulos =', lista_modulos)

    #Crea una lista de OC por cada Modulo de la Auditoria
    matriz_oc=[]
    for mod in lista_modulos:
        
        lista_oc= Obj_Ries.objects.filter(pk_padre=mod.pk)
        matriz_oc.append({'modulo':mod.pk, 'lista_oc':lista_oc})
        #mod.abierto=False   #Cierra las listas de OC de cada modulo en pagina 
        mod.save()  

    #print('matriz oc =', matriz_oc)

    return render(request, 'CIS/Auditorias/mod_aud_list.html',
                  context={'lista_modulos':lista_modulos,'matriz_oc':matriz_oc, 'auditorias':auditorias})


#********************************
#  Crea Modulo en la Auditoria  *
#********************************
from .forms import Crea_ModuloForm

def Crea_Mod(request, pk):

    print('Entra a Crea Modulo Modelo pk=', pk)

    mod = get_object_or_404(Modulo, pk=pk)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    audit=mod.auditoria
    
    # Determinacion de Codigo a asignar.
    cod = mod.nro_submodulos
    mod.nro_submodulos = cod+1
    mod.save()
    
    if cod < 10:
        codigo='0'+str(cod)
    else:
        codigo=str(cod)

    codigo=mod.numero+codigo
    #frame.save()    
    print('codigo =', codigo)

    #Crea Modulo Modelo
    if request.method == 'POST':

        form = Crea_ModuloForm(request.POST)

        if form.is_valid():

            # Crea en registro
            mod_hijo=Modulo()

            mod_hijo.pk_padre=pk
            mod_hijo.pk_aud=mod.pk_aud
            mod_hijo.numero=codigo
            mod_hijo.item=form.cleaned_data['item']
            mod_hijo.nombre=form.cleaned_data['nombre']
            mod_hijo.auditoria=audit

            #Incrementa el nro. de Modulos de la Auditoria
            audit.nro_modulos=audit.nro_modulos+1
            mod_hijo.save()
            audit.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod', args=[str(audit.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_ModuloForm()
        return render(request, 'CIS/Auditorias/crea_modulo_a.html', {'form':form, 'mod':mod, 'audit':audit})
    

#********************************
#  Borra Modulo en la Auditoria *
#********************************

def Borra_Mod(request, pk):

    print('Entra a Borra Modulo en Auditoria pk=', pk)

    mod = get_object_or_404(Modulo, pk=pk)
    mod_p = get_object_or_404(Modulo, pk=mod.pk_padre)

    audit=mod.auditoria

    #Si no hay Objetivos de Control ni Submodulos
    if mod.nro_objetivos == 0 and mod.nro_submodulos <= 0 :
        #Decrementa nro. de modulos en la Auditoria
        audit.nro_modulos = audit.nro_modulos-1

        #Decrementa el nro. de submodulos en el Modulo Padre
        mod_p.nro_submodulos = mod_p.nro_submodulos-2

        audit.save()
        mod_p.save()
        mod.delete()

    return HttpResponseRedirect(reverse('Lista-Mod', args=[str(audit.id)]))



#************************
#  Crea OC la Auditoria *
#************************
from .forms import Crea_OCForm

def Crea_OC(request, pk):

    print('Entra a Crea OC en Auditoria pk=', pk)
 
    modulo = get_object_or_404(Modulo, pk=pk)
    #frame = get_object_or_404(Framework, pk=mod.pk_padre)
    audit=modulo.auditoria
    
    # Determinacion de Codigo a asignar.
    cod = modulo.nro_objetivos
    modulo.nro_objetivos = cod+1
    modulo.save()
    
    if cod < 10:
        codigo='0'+str(cod)
    else:
        codigo=str(cod)

    codigo=modulo.numero+codigo
    #frame.save()    
    print('codigo =', codigo)

    #Crea Modulo Modelo
    if request.method == 'POST':

        form = Crea_OCForm(request.POST)

        if form.is_valid():

            # Crea en registro
            oc=Obj_Ries()

            oc.pk_padre=pk
            oc.numero=codigo
            oc.modulo=modulo
            oc.obj_item=form.cleaned_data['item']
            oc.obj_nombre=form.cleaned_data['nombre']
            oc.obj_descripcion=form.cleaned_data['descripcion']
            oc.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod', args=[str(audit.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_OCForm()
        return render(request, 'CIS/Auditorias/crea_oc_a.html', {'audit':audit, 'modulo':modulo, 'form':form})


#****************************
#  Borra OC de la Auditoria *
#****************************
from .forms import Crea_ModuloForm

def Borra_OC(request, pk):

    print('Entra a Borra Modulo Modelo pk=', pk)

    oc = get_object_or_404(Obj_Ries, pk=pk)
    mod_p = get_object_or_404(Modulo, pk=oc.pk_padre)

    # VAlidar que no tenga Controles ni Comentarios asociados.
    mod_p.nro_objetivos=mod_p.nro_objetivos-1
    oc.delete()
    mod_p.save()

    return HttpResponseRedirect(reverse('Lista-Mod', args=[str(mod_p.pk_aud)]))


#***********************************
#  Modifica el OC de la Auditoria  *
#***********************************
from .forms import Crea_OCForm

def Mod_OC(request, pk):

    print('Entra a Modifica el  OC  pk=', pk)

    oc = get_object_or_404(Obj_Ries, pk=pk)
    mod= get_object_or_404(Modulo, pk=oc.pk_padre)
    audit=get_object_or_404(Auditorias, pk=mod.pk_aud)
        
    #Modifica el  Modulo Modelo
    if request.method == 'POST':

        form = Crea_OCForm(request.POST)

        if form.is_valid():

            # Crea en registro
            
            oc.obj_item=form.cleaned_data['item']
            oc.obj_nombre=form.cleaned_data['nombre']
            oc.obj_descripcion=form.cleaned_data['descripcion']
            oc.save()
            #frame.save()

            #return HttpResponseRedirect(reverse('Lista-Mod-M'))
            return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))

        else:
            print('*** Error de Entrada de Datos *** ')
            print(form.errors)

            #return render(request, 'CIS/mensajes/mensajes_error_Form.html', {'form':form.errors})

    else:

        form = Crea_OCForm(initial={'item':oc.obj_item,
                                 'nombre':oc.obj_nombre,
                                 'descripcion':oc.obj_descripcion})
        
        return render(request, 'CIS/Frameworks/modif_oc_m.html', {'frame':frame, 'mod':mod, 'form':form})


#********************************************************************************
#******************************** Generales  ************************************
#********************************************************************************

#*****************************************
#  Abre o Cierra lista de OC             *
#*****************************************
from .forms import Crea_OCForm

def Abre_Cierra(request, pk):
    """
    Cambia estado del switch para abrir o cerrar la Lista de OCs"""

    print('Entra a Abre/Cierra Mod')

    mod=get_object_or_404(Modulo_Mod, pk=pk)
    frame=mod.framework
    print('Abre =', mod.abierto)

    if mod.abierto:
        mod.abierto=False
        print('Cambia a Falso')
    else:
        mod.abierto=True
        print('cambia a True')

    mod.save()

    return HttpResponseRedirect(reverse('Lista-Mod-M', args=[str(frame.id)]))

def Abre_Cierra_A(request, pk):
    """
    Cambia estado del switch para abrir o cerrar la Lista de OCs"""

    mod=get_object_or_404(Modulo, pk=pk)
    aud=mod.auditoria

    if mod.abierto:
        mod.abierto=False
    else:
        mod.abierto=True

    mod.save()

    return HttpResponseRedirect(reverse('Lista-Mod', args=[str(aud.id)]))

#*****************************************
#  Exporta Framework a  Planilla Excel   *
#*****************************************
# utiliza la biblioteca:  pip install openpyxl

from django.http import HttpResponse
from openpyxl import Workbook



def Export_Frame_to_Exel(request, pk):

    print('***Exporta Framework a Excell***')

    # Recupera los datos de tu modelo (por ejemplo, MiModelo)
    framework=get_object_or_404(Framework, pk=pk)
    obj_control = Obj_Ries_Mod.objects.all()

    # Crea un nuevo libro de trabajo (workbook)
    wb = Workbook()
    ws = wb.active
    ws['A1']='FrameWork :'+framework.fw_nombre+' / '+ framework.fw_version
    ws.title = "OBJ_CTRL"

    # Agrega encabezados de columna
    ws.append(['COD.MODULO','NOMBRE.MODULO','CODIGO.OC', 'NOMBRE.OC', 'DESCRIPCION'])

    # Agrega los datos de tu modelo
    for reg in obj_control:
        modulo=reg.modulo
        frm=int(pk)
        print('frm=', frm)
        frame=reg.modulo.pk_frw*1
        frm1=frm+frame
        print(f"pk: {frm}, frame: {frame}")

        if frame==frm:
            item_modulo=reg.modulo.item
            nombre_modulo=reg.modulo.nombre
            print('datos del modulo=', item_modulo,'-', nombre_modulo)
            ws.append([item_modulo, nombre_modulo, reg.obj_item, reg.obj_nombre, reg.obj_descripcion, ])
        else:
            print('Framework=', frame, '=', pk)
            
            #print('*** sin modulo ***')

    # Crea una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type="CIS/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #response = HttpResponse(content_type="CIS/Downloaded")
    response["Content-Disposition"] = "attachment; filename=framework.xlsx"
    wb.save(response)

    return response

#*****************************************
#  Exporta Auditoria a  Planilla Excel   *
#*****************************************
# utiliza la biblioteca:  pip install openpyxl

from django.http import HttpResponse
from openpyxl import Workbook

def Export_audit_to_Exel(request, pk):

    print('***Exporta Auditoria a Excell***')

    # Recupera los datos de tu modelo (por ejemplo, MiModelo)
    auditoria=get_object_or_404(Auditorias, pk=pk)
    modulos=Modulo.objects.filter(pk_aud=pk)
    #print('pk=', pk , 'pk_aud =', modulos.pk_aud)
    print('modulos:', modulos)

    # Crea un nuevo libro de trabajo (workbook)
    wb = Workbook()
    ws = wb.active
    #ws['A1']='Auditoria :'+auditoria.cod_aud +' - '+ auditoria.nombre
    ws.title = "OBJ_CTRL"

    # Agrega encabezados de columna
    ws.append(['COD.AUD.','NOMBRE.AUD', 'CORR.M', 'COD.MODULO','NOMBRE.MODULO', 'CORR.OC', 'CODIGO.OC', 'NOMBRE.OC', 'DESCRIPCION'])

    # Agrega los datos de tu modelo

    if  modulos:  #si existen modulos
        
        for  mod in modulos:

            obj_control = Obj_Ries.objects.filter(pk_padre=mod.pk)

            if  obj_control:


                for reg in obj_control:
                    
                    pki=int(pk)
                    audit_pk=reg.modulo.pk_aud
                    #print(f"pk: {pki}, audit_pk: {audit_pk}")
                    
                    ws.append([auditoria.cod_aud, auditoria.nombre, mod.numero, mod.item, mod.nombre,
                               reg.numero, reg.obj_item, reg.obj_nombre, reg.obj_descripcion])
                    
            else:
                ws.append([auditoria.cod_aud, auditoria.nombre, mod.numero, mod.item, mod.nombre, '', '', ''])
                    
    else:
        ws.append([auditoria.cod_aud,auditoria.nombre,'Sin Modulos'])
                

    # Crea una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type="CIS/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #response = HttpResponse(content_type="CIS/Downloaded")
    response["Content-Disposition"] = "attachment; filename=auditoria.xlsx"
    wb.save(response)

    return response



#*****************************************************
#  Importar  Planilla Excel al Modulo de Auditoria   *
#*****************************************************
# utiliza la biblioteca:  pip install pandas

from .forms import FileForm

def seleccionar_archivo(request):
    """
    Selecciona planilla a importar y llama al cargador
    """
    print('Entra a seleccionar archivo')

    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        print('vuelve del form')
        if form.is_valid():
            ruta_archivo = form.cleaned_data['ruta']
            print('va hacia funcion de importar')
            
            importar_desde_excel(ruta_archivo)

        else:
            print('ruta no valida')
            print(form.errors)
        
        return HttpResponseRedirect(reverse('Lista-Auditorias'))
    
    else:
        form = FileForm()

    return render(request, 'CIS/Auditorias/imp_auditoria.html', {'form': form})


import os
import pandas as pd
#from .models import Auditorias, Modulo, Obj_Ries  # Asegúrate de importar tus modelos

def importar_desde_excel(ruta_archivo):
    df = pd.read_excel(ruta_archivo)  # Lee la planilla Excel

    corr_mod_ant=''
    corr_oc_ant=''
    fila=0

    for _, row in df.iterrows(): #recorre la Planilla por filas
        fila=fila+1
        print('lee fila :', fila)

        #crea Auditoria
        codigo_aud=row['COD.AUD.']
        nombre_aud=row['NOMBRE.AUD']

        existe_audit=Auditorias.objects.filter(cod_aud=row['COD.AUD.']).exists()
        
        if not existe_audit: 

            # Determinacion de Codigo a asignar.
            cod = get_object_or_404(Parametros_G, pk = 2)
            if cod.valor_2 < 10:
                codigo='0'+str(cod.valor_2)
            else:
                codigo=str(cod.valor_2)
            cod.valor_2=cod.valor_2+1
            cod.save()

            #Crea Audit
            auditoria=Auditorias()

            auditoria.numero=codigo
            auditoria.cod_aud=row['COD.AUD.']
            auditoria.nombre=row['NOMBRE.AUD']
            auditoria.status='P'
            auditoria.save()

            print('Crea Auditoria con el codigo =', auditoria.numero)


        else:
            print('Auditoria ya existe')


        # Carga de Modulos y Objetivos de Control
        #=========================================

       
        audit=get_object_or_404(Auditorias, cod_aud=row['COD.AUD.'])

        if row['COD.MODULO'] == '0.- Root' :

            #Crea Modulo Raiz
            #=================

            modulo_raiz=Modulo()
            modulo_raiz.pk_aud=audit.pk
            modulo_raiz.pk_padre=audit.pk
            modulo_raiz.numero=audit.numero
            modulo_raiz.item=row['COD.MODULO']
            modulo_raiz.nombre=row['NOMBRE.MODULO']
            modulo_raiz.auditoria=audit

            modulo_raiz.save()

            pk_padre=modulo_raiz.pk
            pk_aud=modulo_raiz.pk_aud

            print('Crea el Modulo Raiz modulo.pk=', modulo_raiz.pk)

        else:

            #Crea modulo no raiz
            #====================
            print('CORR.M =', row['CORR.M'],'--', 'cor_mod_ant =', corr_mod_ant,
                  '---', 'COD.MODULO =', row['COD.MODULO'])

            if row['CORR.M'] != corr_mod_ant and row['COD.MODULO'] !='':    #Si existe dato del Modulo y
                                                                            # es distinto al anterior
                                                                            # cree modulo
                modulo=Modulo()

                modulo.pk_aud=pk_aud
                modulo.pk_padre=pk_padre
                modulo.numero=row['CORR.M']                       #Asigna codigo del sistema
                modulo.item=row['COD.MODULO']
                modulo.nombre=row['NOMBRE.MODULO']
                modulo.auditoria=audit

                modulo.save()
                pk_padre=modulo.pk

            #Crea Objetivo de Control
            #========================

            if row['CODIGO.OC'] != '' and row['CORR.OC'] != corr_mod_ant:   #idem OC

                modulo_del_oc=get_object_or_404(Modulo, numero=row['CORR.M'])

                objetivo=Obj_Ries()

                objetivo.pk_padre=modulo_del_oc.pk
                objetivo.numero=row['CORR.OC']
                objetivo.obj_item=row['CODIGO.OC']
                objetivo.obj_nombre=row['NOMBRE.OC']
                objetivo.obj_descripcion=row['DESCRIPCION']
                objetivo.modulo=modulo_del_oc

                objetivo.save()

        corr_mod_ant=row['CORR.M']
        corr_oc_ant=row['CORR.OC']
               





